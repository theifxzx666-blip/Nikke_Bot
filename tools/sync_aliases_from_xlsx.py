# -*- coding: utf-8 -*-
"""从角色卡登记表 xlsx 同步别名 → data/character_aliases_extra.json

用户直接在 nikke_character_dictionary_with_images_updated.xlsx 更新了别名列，
本脚本把"中文名→别名"同步进机器人侧补充别名表（不会被 Codex 覆盖）。
"""
import io
import json
import re
from pathlib import Path

from openpyxl import load_workbook

XLSX = r"F:/Codex/Nikke/Nikke_Wiki/exports/nikke_character_dictionary_with_images_updated.xlsx"
EXTRA_OUT = Path(__file__).resolve().parents[1] / "data" / "character_aliases_extra.json"

# 分隔符：| 和 ；; 混合
SEP = re.compile(r"[|；;]")


def split_aliases(raw: str) -> list[str]:
    parts = [p.strip() for p in SEP.split(raw) if p.strip()]
    return parts


def main():
    wb = load_workbook(XLSX, read_only=True)
    ws = wb["角色字典"]
    result = {}
    skip_self = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        cn = str(row[2] or "").strip() if len(row) > 2 else ""
        raw = str(row[3] or "").strip() if len(row) > 3 else ""
        if not cn or not raw:
            continue
        aliases = split_aliases(raw)
        # 去自身（xlsx 第一个通常是角色自身名）
        real = [a for a in aliases if a and a != cn]
        if real:
            result[cn] = real
        else:
            skip_self += 1

    # 合并：保留旧补充表中 xlsx 未覆盖的角色条目（如 艾达→艾达王/王姐）
    old = {}
    if EXTRA_OUT.exists():
        with io.open(EXTRA_OUT, encoding="utf-8-sig") as f:
            old = json.load(f)
    merged = dict(old)
    for cn, aliases in result.items():
        merged[cn] = aliases  # xlsx 为准（用户最新意图）
    # 从 xlsx 删除条目的角色（xlsx 有行但无别名）→ 从补充表移除
    # 识别 xlsx 中出现的所有中文名
    xlsx_names = set()
    wb2 = load_workbook(XLSX, read_only=True)
    ws2 = wb2["角色字典"]
    for row in ws2.iter_rows(min_row=2, values_only=True):
        cn = str(row[2] or "").strip() if len(row) > 2 else ""
        if cn:
            xlsx_names.add(cn)
    removed = [k for k in list(merged.keys()) if k in xlsx_names and k not in result]
    for k in removed:
        del merged[k]
    if removed:
        print(f"移除（xlsx 该角色已无别名）: {removed}")

    print(f"同步角色数: {len(result)}（跳过纯自引用 {skip_self} 行），合并后共 {len(merged)} 角色")

    with io.open(EXTRA_OUT, "w", encoding="utf-8", newline="\n") as f:
        json.dump(merged, f, ensure_ascii=False, indent=2)
    print(f"已写入 {EXTRA_OUT}")

    # 预览几条
    for cn, aliases in list(result.items())[:8]:
        print(f"  {cn}: {aliases}")


if __name__ == "__main__":
    main()
