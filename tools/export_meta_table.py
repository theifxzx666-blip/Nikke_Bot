# -*- coding: utf-8 -*-
"""导出培养建议核对表（Excel）：角色名 + 别名 + 全部培养字段。

用法：
    PYTHONPATH=F:/Codex/Nikke/Nikke_Bot .venv/Scripts/python.exe tools/export_meta_table.py
输出：
    F:/Codex/Nikke/Nikke_Bot/data/培养建议核对表.xlsx
"""
import io
import json
import os
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from guild_war_bot.wiki_query import default_index

BASE_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BASE_DIR / "data"
META_OUT = DATA_DIR / "character_meta.json"
EXTRA_ALIASES = DATA_DIR / "character_aliases_extra.json"
XLSX_OUT = DATA_DIR / "培养建议核对表.xlsx"

# 手工修正（与 build_character_meta.py 的 MANUAL_FIXES 保持一致）
MANUAL_FIXED = {"灰姑娘", "红莲：暗影", "格拉维", "桃乐丝", "索达：闪亮兔女郎"}


def load_aliases() -> dict[str, list[str]]:
    """聚合所有别名来源 → {中文正式名: [别名...]}"""
    idx = default_index()
    result = defaultdict(list)
    # 1. WikiIndex by_alias（nikke_character_aliases.json + extra 合并）
    #    formal 可能是英文名（Ada）或中文名，统一解析到 cnName
    for alias, formal in idx.by_alias.items():
        rec = idx.by_cnname.get(formal) or idx.by_name.get(formal)
        cn = str(rec.get("cnName") or "").strip() if rec else str(formal).strip()
        if not cn:
            continue
        if alias == cn:
            continue  # 跳过自引用（"丽塔"→丽塔 无信息量）
        if alias not in result[cn]:
            result[cn].append(alias)
    # 2. 补充词典 characters_extra.json 的内嵌 aliases
    extra_chars = DATA_DIR / "characters_extra.json"
    if extra_chars.exists():
        with io.open(extra_chars, encoding="utf-8-sig") as f:
            for rec in json.load(f):
                cn = str(rec.get("cnName") or "").strip()
                al = str(rec.get("aliases") or "").strip()
                if cn and al:
                    for a in al.split("|"):
                        a = a.strip()
                        if a and a not in result[cn]:
                            result[cn].append(a)
    return dict(result)


def main():
    # 培养数据
    with io.open(META_OUT, encoding="utf-8-sig") as f:
        meta = json.load(f)
    aliases = load_aliases()

    wb = Workbook()
    ws = wb.active
    ws.title = "培养建议"

    headers = ["角色名", "别名", "强度", "装备", "词条", "备注", "技能", "魔方", "收藏品", "校对状态"]
    ws.append(headers)

    # 表头样式
    head_fill = PatternFill("solid", fgColor="4472C4")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for name, rec in sorted(meta.items()):
        if name.startswith("_"):
            continue
        aliases_str = "、".join(aliases.get(name, [])) if aliases.get(name) else ""
        status = "✅ 已校对" if name in MANUAL_FIXED else "待校对"
        ws.append([
            name,
            aliases_str,
            rec.get("strength", ""),
            rec.get("gear", ""),
            rec.get("gear_stat", ""),
            rec.get("note", ""),
            rec.get("skill", ""),
            rec.get("cube", ""),
            rec.get("collection", ""),
            status,
        ])

    # 列宽
    widths = [18, 40, 12, 22, 30, 40, 16, 28, 14, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 对齐（备注/词条自动换行）
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # 冻结首行 + 筛选
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(XLSX_OUT)
    print(f"✅ 已导出: {XLSX_OUT}")
    print(f"   角色数: {len(meta)} | 列数: {len(headers)}")


if __name__ == "__main__":
    main()
