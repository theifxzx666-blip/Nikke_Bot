from __future__ import annotations

import os
import re
import shutil
import sqlite3
import subprocess
import sys
import threading
import uuid
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .core import GuildWarBot, MAX_ATTACKS_PER_MEMBER, RAID_START, battle_day


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DB_PATH = ROOT / "data" / "guild_war.db"
DEFAULT_SAMPLE_DB_PATH = ROOT / "data" / "union_sample.db"
DEFAULT_OUTPUT_DIR = ROOT / "data" / "exports"
DEFAULT_TEMPLATE_PATH = Path(os.environ.get(
    "UNION_RAID_TEMPLATE",
    r"D:\Codex\依赖\嗷润吉-联盟突袭统计.xlsx",
))
FLOW1_SETTING_DEFAULTS = {
    "flow1_raid_day": "1",
    "flow1_battle_date": "",
    "flow1_pages": "12",
    "flow1_scroll_rows": "6",
    "flow1_drag_start_row": "6",
    "flow1_drag_distance_rows": "6.1",
    "flow1_drag_end_safe_ratio": "0.16",
    "flow1_use_drag_anchor": "1",
    "flow1_drag_anchor_start": "0.53,0.79",
    "flow1_drag_anchor_end": "0.53,0.25",
    "flow1_union_point": "0.912,0.395",
    "flow1_raid_entry_point": "0.50,0.835",
    "flow1_record_point": "",
    "flow1_day_tab_point": "0.34,0.225",
    "flow1_day1_point": "0.16,0.278",
    "flow1_day2_point": "0.31,0.278",
    "flow1_drag_duration_seconds": "1.4",
    "flow1_drag_steps": "56",
    "flow1_drag_hold_seconds": "0.35",
    "flow1_skip_open_record": "0",
}
FLOW1_DEBUG_DIR = ROOT / "data" / "flow1_debug"


@dataclass(frozen=True)
class BossSnapshot:
    id: int
    captured_at: str
    raid_day: int | None
    boss_index: int | None
    boss_label: str
    boss_name: str
    level: int | None
    current_hp: int | None
    total_hp: int | None
    percent: float | None
    image_path: str
    note: str


@dataclass(frozen=True)
class TaskRun:
    id: str
    task_type: str
    title: str
    status: str
    command: str
    started_at: str
    finished_at: str
    output: str
    artifact_path: str


class UnionRaidSuite:
    """Unified data layer for the local union raid management program."""

    def __init__(
        self,
        db_path: Path | str = DEFAULT_DB_PATH,
        sample_db_path: Path | str = DEFAULT_SAMPLE_DB_PATH,
    ) -> None:
        self.db_path = Path(db_path)
        self.sample_db_path = Path(sample_db_path)
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self.bot = GuildWarBot(self.db_path)
        self.conn = self.bot.conn
        self.lock = self.bot.lock
        self.init_db()

    def close(self) -> None:
        self.bot.close()

    def init_db(self) -> None:
        with self.lock:
            for column, column_type in {
                "boss_label": "text",
                "boss_name": "text",
                "team_text": "text",
                "raid_day": "integer",
                "is_tail": "integer not null default 0",
                "source": "text",
            }.items():
                self.bot.ensure_column("attacks", column, column_type)
            self.conn.executescript(
                """
                create table if not exists boss_snapshots (
                    id integer primary key autoincrement,
                    captured_at text not null,
                    raid_day integer,
                    boss_index integer,
                    boss_label text not null default '',
                    boss_name text not null default '',
                    level integer,
                    current_hp integer,
                    total_hp integer,
                    percent real,
                    image_path text not null default '',
                    note text not null default ''
                );

                create table if not exists task_runs (
                    id text primary key,
                    task_type text not null,
                    title text not null,
                    status text not null,
                    command text not null,
                    started_at text not null,
                    finished_at text not null default '',
                    output text not null default '',
                    artifact_path text not null default ''
                );

                create table if not exists app_settings (
                    key text primary key,
                    value text not null default ''
                );
                """
            )
            self.conn.commit()

    def dashboard(self) -> dict[str, Any]:
        day = battle_day()
        members = self.bot.list_member_records(include_inactive=True)
        active_members = [member for member in members if member.active]
        statuses = self.bot.statuses(day)
        total_attacks = sum(status.attacks for status in statuses)
        total_required = len(active_members) * MAX_ATTACKS_PER_MEMBER
        latest_snapshot = self.latest_boss_snapshot()
        latest_task = self.list_tasks(limit=1)
        return {
            "battle_date": day.isoformat(),
            "members_total": len(members),
            "members_active": len(active_members),
            "attacks_done": total_attacks,
            "attacks_required": total_required,
            "latest_snapshot": latest_snapshot,
            "latest_task": latest_task[0] if latest_task else None,
            "sample_count": self.sample_member_count(),
        }

    def get_setting(self, key: str, default: str = "") -> str:
        with self.lock:
            row = self.conn.execute(
                "select value from app_settings where key = ?",
                (key,),
            ).fetchone()
        return str(row["value"]) if row else default

    def set_setting(self, key: str, value: str) -> None:
        with self.lock:
            self.conn.execute(
                """
                insert into app_settings(key, value)
                values(?, ?)
                on conflict(key) do update set value = excluded.value
                """,
                (key, value),
            )
            self.conn.commit()

    def flow1_settings(self) -> dict[str, str]:
        settings = dict(FLOW1_SETTING_DEFAULTS)
        with self.lock:
            rows = self.conn.execute(
                "select key, value from app_settings where key like 'flow1_%'"
            ).fetchall()
        for row in rows:
            key = str(row["key"])
            if key in settings:
                settings[key] = str(row["value"] or "")
        if not settings["flow1_battle_date"]:
            settings["flow1_battle_date"] = battle_day().isoformat()
        return settings

    def save_flow1_settings(self, values: dict[str, str]) -> str:
        clean = dict(FLOW1_SETTING_DEFAULTS)
        clean.update({key: str(values.get(key) or "").strip() for key in clean})
        if not clean["flow1_raid_day"]:
            clean["flow1_raid_day"] = FLOW1_SETTING_DEFAULTS["flow1_raid_day"]
        if not clean["flow1_battle_date"]:
            clean["flow1_battle_date"] = battle_day().isoformat()
        if not clean["flow1_pages"]:
            clean["flow1_pages"] = FLOW1_SETTING_DEFAULTS["flow1_pages"]
        if not clean["flow1_scroll_rows"]:
            clean["flow1_scroll_rows"] = FLOW1_SETTING_DEFAULTS["flow1_scroll_rows"]
        if not clean["flow1_drag_start_row"]:
            clean["flow1_drag_start_row"] = FLOW1_SETTING_DEFAULTS["flow1_drag_start_row"]
        if not clean["flow1_drag_distance_rows"]:
            clean["flow1_drag_distance_rows"] = FLOW1_SETTING_DEFAULTS["flow1_drag_distance_rows"]
        if not clean["flow1_drag_end_safe_ratio"]:
            clean["flow1_drag_end_safe_ratio"] = FLOW1_SETTING_DEFAULTS["flow1_drag_end_safe_ratio"]
        clean["flow1_use_drag_anchor"] = "1" if clean["flow1_use_drag_anchor"] == "1" else "0"
        if not clean["flow1_drag_anchor_start"]:
            clean["flow1_drag_anchor_start"] = FLOW1_SETTING_DEFAULTS["flow1_drag_anchor_start"]
        if not clean["flow1_drag_anchor_end"]:
            clean["flow1_drag_anchor_end"] = FLOW1_SETTING_DEFAULTS["flow1_drag_anchor_end"]
        for key in (
            "flow1_union_point",
            "flow1_raid_entry_point",
            "flow1_day_tab_point",
            "flow1_day1_point",
            "flow1_day2_point",
            "flow1_drag_duration_seconds",
            "flow1_drag_steps",
            "flow1_drag_hold_seconds",
        ):
            if not clean[key]:
                clean[key] = FLOW1_SETTING_DEFAULTS[key]
        clean["flow1_skip_open_record"] = "1" if clean["flow1_skip_open_record"] == "1" else "0"
        with self.lock:
            self.conn.executemany(
                """
                insert into app_settings(key, value)
                values(?, ?)
                on conflict(key) do update set value = excluded.value
                """,
                sorted(clean.items()),
            )
            self.conn.commit()
        return "流程1采样参数已保存。"

    def latest_flow1_artifacts(self) -> list[dict[str, str]]:
        latest_paths = FLOW1_DEBUG_DIR / "latest_paths.txt"
        parsed: dict[str, str] = {}
        if latest_paths.exists():
            for line in latest_paths.read_text(encoding="utf-8", errors="replace").splitlines():
                if "=" not in line:
                    continue
                key, value = line.split("=", 1)
                parsed[key.strip()] = value.strip()

        items = [
            ("识别明细 CSV", FLOW1_DEBUG_DIR / "latest_records.csv"),
            ("出刀数 CSV", FLOW1_DEBUG_DIR / "latest_attendance.csv"),
            ("识别原始 JSON", FLOW1_DEBUG_DIR / "latest_records.json"),
            ("操作日志", FLOW1_DEBUG_DIR / "latest_operation.log"),
            ("最新路径说明", latest_paths),
        ]
        artifacts: list[dict[str, str]] = []
        for label, path in items:
            artifacts.append(
                {
                    "label": label,
                    "path": str(path.resolve()),
                    "exists": "1" if path.exists() and path.is_file() else "0",
                }
            )
        session_dir = parsed.get("session_dir", "")
        if session_dir:
            artifacts.append(
                {
                    "label": "最新运行目录",
                    "path": session_dir,
                    "exists": "1" if Path(session_dir).exists() else "0",
                }
            )
        return artifacts

    def sample_member_count(self) -> int:
        if not self.sample_db_path.exists():
            return 0
        try:
            conn = sqlite3.connect(self.sample_db_path)
            row = conn.execute("select count(*) from member_samples").fetchone()
            conn.close()
            return int(row[0]) if row else 0
        except sqlite3.Error:
            return 0

    def latest_sample_members(self, limit: int = 32) -> list[dict[str, Any]]:
        if not self.sample_db_path.exists():
            return []
        conn = sqlite3.connect(self.sample_db_path)
        conn.row_factory = sqlite3.Row
        try:
            rows = conn.execute(
                """
                select name, power, level, online_text, online_status, updated_at
                from member_samples
                where trim(name) <> ''
                order by
                    case when power is null then 1 else 0 end,
                    power desc,
                    updated_at desc,
                    name
                limit ?
                """,
                (limit,),
            ).fetchall()
            return [dict(row) for row in rows]
        finally:
            conn.close()

    def import_sampled_members(self, limit: int = 32, server_area: str | None = None) -> str:
        rows = self.latest_sample_members(limit)
        if not rows:
            return "没有找到队员采样数据。请先运行队员采样。"
        added = 0
        skipped = 0
        for row in rows:
            name = str(row.get("name") or "").strip()
            if not name:
                skipped += 1
                continue
            before = self.bot.member_count(include_inactive=True)
            self.bot.add_member(name, server_area=server_area)
            after = self.bot.member_count(include_inactive=True)
            if after > before:
                added += 1
            else:
                skipped += 1
        return f"采样成员同步完成：新增 {added} 人，跳过 {skipped} 人。"

    def add_member(
        self,
        name: str,
        server_area: str | None,
        qq: str | None,
        group_card: str | None,
    ) -> str:
        return self.bot.add_member(name, qq=qq, group_card=group_card, server_area=server_area)

    def update_members(self, rows: list[dict[str, str]]) -> str:
        return self.bot.bulk_update_members(rows)

    def list_members(self) -> list[Any]:
        return self.bot.list_member_records(include_inactive=True)

    def add_boss_snapshot(
        self,
        raid_day: int | None,
        boss_index: int | None,
        boss_label: str,
        boss_name: str,
        level: int | None,
        current_hp: int | None,
        total_hp: int | None,
        image_path: str = "",
        note: str = "",
        captured_at: str | None = None,
    ) -> str:
        percent = None
        if current_hp is not None and total_hp:
            percent = current_hp / total_hp
        with self.lock:
            self.conn.execute(
                """
                insert into boss_snapshots(
                    captured_at, raid_day, boss_index, boss_label, boss_name,
                    level, current_hp, total_hp, percent, image_path, note
                )
                values(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    captured_at or now_text(),
                    raid_day,
                    boss_index,
                    boss_label.strip(),
                    boss_name.strip(),
                    level,
                    current_hp,
                    total_hp,
                    percent,
                    image_path.strip(),
                    note.strip(),
                ),
            )
            self.conn.commit()
        return "Boss 进度已保存。"

    def latest_boss_snapshot(self) -> BossSnapshot | None:
        with self.lock:
            row = self.conn.execute(
                "select * from boss_snapshots order by captured_at desc, id desc limit 1"
            ).fetchone()
        return row_to_snapshot(row) if row else None

    def list_boss_snapshots(self, limit: int = 30) -> list[BossSnapshot]:
        with self.lock:
            rows = self.conn.execute(
                "select * from boss_snapshots order by captured_at desc, id desc limit ?",
                (limit,),
            ).fetchall()
        return [row_to_snapshot(row) for row in rows]

    def add_attack(
        self,
        member_id: int,
        battle_date_text: str,
        damage: int | None,
        note: str,
        boss_label: str,
        boss_name: str,
        team_text: str,
        raid_day: int | None,
        is_tail: bool,
        source: str = "manual",
    ) -> str:
        day = parse_date_text(battle_date_text) or battle_day()
        with self.lock:
            member = self.conn.execute(
                "select id, name from members where id = ? and active = 1",
                (member_id,),
            ).fetchone()
            if not member:
                return "找不到启用状态的成员。"
            count = self.bot.attack_count(member_id, day.isoformat())
            if count >= MAX_ATTACKS_PER_MEMBER:
                return f"{member['name']} 在 {day.isoformat()} 已经记录满 {MAX_ATTACKS_PER_MEMBER} 刀。"
            self.conn.execute(
                """
                insert into attacks(
                    member_id, battle_date, damage, note, boss_label, boss_name,
                    team_text, raid_day, is_tail, source
                )
                values(?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    member_id,
                    day.isoformat(),
                    damage,
                    note.strip(),
                    boss_label.strip(),
                    boss_name.strip(),
                    team_text.strip(),
                    raid_day,
                    1 if is_tail else 0,
                    source,
                ),
            )
            self.conn.commit()
        return f"已记录：{member['name']} 第 {count + 1}/{MAX_ATTACKS_PER_MEMBER} 刀。"

    def import_flow1_records(
        self,
        records: list[dict[str, Any]],
        source: str = "flow1_sampler",
        replace: bool = True,
    ) -> dict[str, int]:
        """Import flow1 day-record sampler rows without treating D1 shortfall as an error."""
        inserted = 0
        skipped_unmatched = 0
        skipped_damage = 0
        skipped_limit = 0
        skipped_duplicate = 0
        if not records:
            return {
                "inserted": 0,
                "skipped_unmatched": 0,
                "skipped_damage": 0,
                "skipped_limit": 0,
                "skipped_duplicate": 0,
            }

        first_day = parse_date_text(str(records[0].get("battle_date") or "")) or battle_day()
        first_raid_day = records[0].get("raid_day")
        raid_day_value = int(first_raid_day) if first_raid_day not in (None, "") else None
        clean_source = (source or "flow1_sampler").strip()

        with self.lock:
            if replace:
                if raid_day_value is None:
                    self.conn.execute(
                        "delete from attacks where battle_date = ? and source = ? and raid_day is null",
                        (first_day.isoformat(), clean_source),
                    )
                else:
                    self.conn.execute(
                        "delete from attacks where battle_date = ? and source = ? and raid_day = ?",
                        (first_day.isoformat(), clean_source, raid_day_value),
                    )

            counters: dict[int, int] = {}
            existing_rows = self.conn.execute(
                """
                select member_id, count(*) as attacks
                from attacks
                where battle_date = ?
                group by member_id
                """,
                (first_day.isoformat(),),
            ).fetchall()
            for row in existing_rows:
                counters[int(row["member_id"])] = int(row["attacks"])

            seen: set[tuple[Any, ...]] = set()
            for record in records:
                member_id = record.get("member_id")
                try:
                    member_id = int(member_id)
                except (TypeError, ValueError):
                    member_id = None
                if not member_id:
                    skipped_unmatched += 1
                    continue
                member = self.conn.execute(
                    "select id, name from members where id = ? and active = 1",
                    (member_id,),
                ).fetchone()
                if not member:
                    skipped_unmatched += 1
                    continue
                damage = record.get("damage")
                try:
                    damage = int(damage)
                except (TypeError, ValueError):
                    skipped_damage += 1
                    continue
                duplicate_key = (
                    member_id,
                    damage,
                    str(record.get("boss_name") or ""),
                    record.get("boss_level"),
                    str(record.get("row_hash") or ""),
                )
                if duplicate_key in seen:
                    skipped_duplicate += 1
                    continue
                seen.add(duplicate_key)
                if counters.get(member_id, 0) >= MAX_ATTACKS_PER_MEMBER:
                    skipped_limit += 1
                    continue
                note_parts = [
                    str(record.get("member_raw") or "").strip(),
                    str(record.get("damage_method") or "").strip(),
                    str(record.get("row_image_path") or "").strip(),
                ]
                note = "flow1 " + " | ".join(part for part in note_parts if part)
                self.conn.execute(
                    """
                    insert into attacks(
                        member_id, battle_date, damage, note, boss_label, boss_name,
                        team_text, raid_day, is_tail, source
                    )
                    values(?, ?, ?, ?, ?, ?, '', ?, 0, ?)
                    """,
                    (
                        member_id,
                        first_day.isoformat(),
                        damage,
                        note[:1000],
                        str(record.get("boss_label") or "").strip(),
                        str(record.get("boss_name") or "").strip(),
                        raid_day_value,
                        clean_source,
                    ),
                )
                counters[member_id] = counters.get(member_id, 0) + 1
                inserted += 1
            self.conn.commit()

        return {
            "inserted": inserted,
            "skipped_unmatched": skipped_unmatched,
            "skipped_damage": skipped_damage,
            "skipped_limit": skipped_limit,
            "skipped_duplicate": skipped_duplicate,
        }

    def delete_attack(self, attack_id: int) -> str:
        with self.lock:
            row = self.conn.execute(
                "select id from attacks where id = ?",
                (attack_id,),
            ).fetchone()
            if not row:
                return "找不到这条出刀记录。"
            self.conn.execute("delete from attacks where id = ?", (attack_id,))
            self.conn.commit()
        return "出刀记录已删除。"

    def list_attacks(
        self,
        limit: int = 120,
        battle_date: date | str | None = None,
        newest_first: bool = True,
    ) -> list[dict[str, Any]]:
        day = normalize_date_value(battle_date)
        where = ""
        params: tuple[Any, ...] = ()
        if day:
            where = "where a.battle_date = ?"
            params = (day,)
        with self.lock:
            rows = self.conn.execute(
                f"""
                select
                    a.*,
                    m.name as member_name,
                    m.qq as member_qq,
                    m.group_card as group_card
                from attacks a
                join members m on m.id = a.member_id
                {where}
                order by a.battle_date asc, a.member_id asc, a.id asc
                """,
                params,
            ).fetchall()

        result: list[dict[str, Any]] = []
        counters: dict[tuple[str, int], int] = {}
        for row in rows:
            key = (str(row["battle_date"]), int(row["member_id"]))
            counters[key] = counters.get(key, 0) + 1
            item = dict(row)
            item["attack_no"] = counters[key]
            result.append(item)
        if newest_first:
            result = list(reversed(result))
        return result[:limit] if limit > 0 else result

    def attendance_rows_for_date(self, battle_date: date | str | None = None) -> list[dict[str, Any]]:
        day = normalize_date_value(battle_date) or battle_day().isoformat()
        members = self.bot.list_member_records(include_inactive=False)
        with self.lock:
            raw = self.conn.execute(
                """
                select
                    a.*,
                    m.name as member_name,
                    m.qq as member_qq,
                    m.group_card as group_card
                from attacks a
                join members m on m.id = a.member_id
                where a.battle_date = ?
                order by a.member_id asc, a.id asc
                """,
                (day,),
            ).fetchall()
        by_member: dict[int, list[dict[str, Any]]] = {}
        for row in raw:
            item = dict(row)
            by_member.setdefault(int(row["member_id"]), []).append(item)

        rows: list[dict[str, Any]] = []
        total_damage = sum(int(row["damage"] or 0) for row in raw)
        for member in members:
            attacks = by_member.get(member.id, [])
            damages = [int(attack.get("damage") or 0) for attack in attacks[:MAX_ATTACKS_PER_MEMBER]]
            while len(damages) < MAX_ATTACKS_PER_MEMBER:
                damages.append(0)
            notes = []
            for index, attack in enumerate(attacks[:MAX_ATTACKS_PER_MEMBER], start=1):
                boss = " ".join(
                    str(attack.get(key) or "").strip()
                    for key in ("boss_label", "boss_name")
                ).strip()
                team = str(attack.get("team_text") or "").strip()
                tail = "尾刀" if int(attack.get("is_tail") or 0) else ""
                pieces = [piece for piece in (boss, team, tail) if piece]
                if pieces:
                    notes.append(f"第{index}刀:" + " ".join(pieces))
            total = sum(damages)
            rows.append(
                {
                    "battle_date": day,
                    "member": member,
                    "attack_count": len(attacks),
                    "damages": damages,
                    "total_damage": total,
                    "damage_share": total / total_damage if total_damage else 0,
                    "remaining": max(0, MAX_ATTACKS_PER_MEMBER - len(attacks)),
                    "notes": "；".join(notes),
                    "attacks": attacks,
                }
            )
        return rows

    def attendance_rows(self) -> list[dict[str, Any]]:
        members = self.bot.list_member_records(include_inactive=False)
        day_dates = [raid_day_date(1), raid_day_date(2)]
        with self.lock:
            raw = self.conn.execute(
                """
                select member_id, battle_date, count(*) as attacks, coalesce(sum(damage), 0) as damage
                from attacks
                group by member_id, battle_date
                """
            ).fetchall()
        by_member_day: dict[tuple[int, str], sqlite3.Row] = {
            (int(row["member_id"]), str(row["battle_date"])): row for row in raw
        }
        rows: list[dict[str, Any]] = []
        for member in members:
            d1_row = by_member_day.get((member.id, day_dates[0].isoformat()))
            d2_row = by_member_day.get((member.id, day_dates[1].isoformat()))
            d1 = int(d1_row["attacks"]) if d1_row else 0
            d2 = int(d2_row["attacks"]) if d2_row else 0
            total = d1 + d2
            rows.append(
                {
                    "member": member,
                    "d1": d1,
                    "d2": d2,
                    "total": total,
                    "remaining": max(0, MAX_ATTACKS_PER_MEMBER * 2 - total),
                }
            )
        return rows

    def list_tasks(self, limit: int = 12) -> list[TaskRun]:
        with self.lock:
            rows = self.conn.execute(
                "select * from task_runs order by started_at desc limit ?",
                (limit,),
            ).fetchall()
        return [row_to_task(row) for row in rows]

    def start_command_task(
        self,
        task_type: str,
        title: str,
        args: list[str],
        cwd: Path | str = ROOT,
    ) -> str:
        task_id = uuid.uuid4().hex[:12]
        command = " ".join(quote_arg(arg) for arg in args)
        with self.lock:
            self.conn.execute(
                """
                insert into task_runs(id, task_type, title, status, command, started_at)
                values(?, ?, ?, 'running', ?, ?)
                """,
                (task_id, task_type, title, command, now_text()),
            )
            self.conn.commit()

        thread = threading.Thread(
            target=self._run_task,
            args=(task_id, task_type, args, Path(cwd)),
            name=f"union-task-{task_id}",
            daemon=True,
        )
        thread.start()
        return f"任务已启动：{title}。刷新页面可查看结果。"

    def _run_task(self, task_id: str, task_type: str, args: list[str], cwd: Path) -> None:
        status = "success"
        output = ""
        artifact_path = ""
        try:
            env = os.environ.copy()
            env.setdefault("PYTHONUTF8", "1")
            completed = subprocess.run(
                args,
                cwd=str(cwd),
                env=env,
                text=True,
                encoding="utf-8",
                errors="replace",
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                check=False,
            )
            output = completed.stdout or ""
            status = "success" if completed.returncode == 0 else "failed"
            artifact_path = find_existing_artifact(output)
            if task_type == "progress_capture" and artifact_path:
                self.add_boss_snapshot(
                    raid_day=None,
                    boss_index=None,
                    boss_label="",
                    boss_name="",
                    level=None,
                    current_hp=None,
                    total_hp=None,
                    image_path=artifact_path,
                    note="自动截图已保存，血量数据待校对。",
                )
        except Exception as exc:
            status = "failed"
            output = f"{type(exc).__name__}: {exc}"
        with self.lock:
            self.conn.execute(
                """
                update task_runs
                set status = ?, finished_at = ?, output = ?, artifact_path = ?
                where id = ?
                """,
                (status, now_text(), output[-12000:], artifact_path, task_id),
            )
            self.conn.commit()

    def export_workbook(
        self,
        template_path: Path | str = DEFAULT_TEMPLATE_PATH,
        output_dir: Path | str = DEFAULT_OUTPUT_DIR,
    ) -> Path:
        output_root = Path(output_dir)
        output_root.mkdir(parents=True, exist_ok=True)
        template = Path(template_path)
        template_note = ""
        if template.exists():
            try:
                workbook = load_workbook(template)
            except Exception as exc:
                workbook = Workbook()
                template_note = (
                    f"模板读取失败，已生成独立工作簿。模板：{template}；"
                    f"错误：{type(exc).__name__}: {exc}"
                )
        else:
            workbook = Workbook()
            template_note = f"未找到模板，已生成独立工作簿。模板：{template}"
        for name in ["当前-考勤", "当前-统伤", "出刀明细", "Boss快照", "成员绑定"]:
            if name in workbook.sheetnames:
                del workbook[name]

        if template_note:
            build_note_sheet(workbook.create_sheet("导出说明"), template_note)
        build_attendance_sheet(workbook.create_sheet("当前-考勤"), self.attendance_rows())
        build_damage_sheet(workbook.create_sheet("当前-统伤"), self.list_attacks(limit=1000))
        build_attack_detail_sheet(workbook.create_sheet("出刀明细"), self.list_attacks(limit=2000))
        build_snapshot_sheet(workbook.create_sheet("Boss快照"), self.list_boss_snapshots(limit=200))
        build_member_sheet(workbook.create_sheet("成员绑定"), self.list_members())

        if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
            sheet = workbook["Sheet"]
            if sheet.max_row == 1 and sheet.max_column == 1 and sheet["A1"].value is None:
                del workbook["Sheet"]

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = output_root / f"联盟突袭管理导出_{timestamp}.xlsx"
        workbook.save(output_path)
        return output_path.resolve()

    def export_day_workbook(
        self,
        battle_date: date | str | None = None,
        output_dir: Path | str = DEFAULT_OUTPUT_DIR,
    ) -> Path:
        day = normalize_date_value(battle_date) or battle_day().isoformat()
        output_root = Path(output_dir)
        output_root.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        default_sheet = workbook.active
        default_sheet.title = f"{day}-出刀数"
        day_rows = self.attendance_rows_for_date(day)
        attacks = self.list_attacks(limit=0, battle_date=day, newest_first=True)

        build_day_attendance_sheet(default_sheet, day_rows)
        damage_sheet = workbook.create_sheet(f"{day}-统伤")
        build_reference_damage_sheet(damage_sheet, attacks, day)
        build_attack_detail_sheet(workbook.create_sheet(f"{day}-出刀明细"), attacks)
        build_member_sheet(workbook.create_sheet("成员绑定"), self.list_members())

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = output_root / f"联盟突袭_{day}_出刀明细_{timestamp}.xlsx"
        workbook.save(output_path)
        return output_path.resolve()


def row_to_snapshot(row: sqlite3.Row) -> BossSnapshot:
    return BossSnapshot(
        id=int(row["id"]),
        captured_at=str(row["captured_at"] or ""),
        raid_day=int(row["raid_day"]) if row["raid_day"] is not None else None,
        boss_index=int(row["boss_index"]) if row["boss_index"] is not None else None,
        boss_label=str(row["boss_label"] or ""),
        boss_name=str(row["boss_name"] or ""),
        level=int(row["level"]) if row["level"] is not None else None,
        current_hp=int(row["current_hp"]) if row["current_hp"] is not None else None,
        total_hp=int(row["total_hp"]) if row["total_hp"] is not None else None,
        percent=float(row["percent"]) if row["percent"] is not None else None,
        image_path=str(row["image_path"] or ""),
        note=str(row["note"] or ""),
    )


def row_to_task(row: sqlite3.Row) -> TaskRun:
    return TaskRun(
        id=str(row["id"]),
        task_type=str(row["task_type"] or ""),
        title=str(row["title"] or ""),
        status=str(row["status"] or ""),
        command=str(row["command"] or ""),
        started_at=str(row["started_at"] or ""),
        finished_at=str(row["finished_at"] or ""),
        output=str(row["output"] or ""),
        artifact_path=str(row["artifact_path"] or ""),
    )


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def parse_int_text(raw: str | None) -> int | None:
    text = (raw or "").strip().replace(",", "")
    if not text:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    value = float(match.group(0))
    if "亿" in text:
        value *= 100_000_000
    elif "万" in text or text.lower().endswith("w"):
        value *= 10_000
    return int(value)


def parse_date_text(raw: str | None) -> date | None:
    text = (raw or "").strip()
    if not text:
        return None
    try:
        return date.fromisoformat(text)
    except ValueError:
        return None


def normalize_date_value(raw: date | str | None) -> str | None:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date().isoformat()
    if isinstance(raw, date):
        return raw.isoformat()
    parsed = parse_date_text(str(raw))
    return parsed.isoformat() if parsed else None


def raid_day_date(day_number: int) -> date:
    return RAID_START.date() + timedelta(days=max(0, day_number - 1))


def format_number(value: int | float | None) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.1%}"
    return f"{value:,}"


def format_hp(current_hp: int | None, total_hp: int | None, percent: float | None) -> str:
    if current_hp is None and total_hp is None:
        return ""
    hp = f"{format_number(current_hp)}/{format_number(total_hp)}"
    if percent is not None:
        hp += f" ({percent:.1%})"
    return hp


def find_existing_artifact(output: str) -> str:
    patterns = [
        r"[A-Za-z]:[\\/][^\r\n\"<>|]+?\.(?:png|jpg|jpeg|xlsx|csv)",
        r"[A-Za-z]:[\\/][^\r\n\"<>|]+",
    ]
    for pattern in patterns:
        for match in reversed(re.findall(pattern, output)):
            path = Path(match.strip())
            if path.exists() and path.is_file():
                return str(path.resolve())
    return ""


def quote_arg(value: str) -> str:
    if re.search(r"\s", value):
        return f'"{value}"'
    return value


def style_header(sheet: Any) -> None:
    style_header_row(sheet, 1)


def style_header_row(sheet: Any, row_number: int) -> None:
    fill = PatternFill("solid", fgColor="EAF1FB")
    for cell in sheet[row_number]:
        cell.font = Font(bold=True, color="1F2937")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = f"A{row_number + 1}"


def autosize(sheet: Any, max_width: int = 34) -> None:
    for column in sheet.columns:
        width = 10
        letter = get_column_letter(column[0].column)
        for cell in column:
            text = "" if cell.value is None else str(cell.value)
            width = max(width, min(max_width, len(text) + 2))
        sheet.column_dimensions[letter].width = width


def build_attendance_sheet(sheet: Any, rows: list[dict[str, Any]]) -> None:
    headers = ["成员ID", "区服", "D1", "D2", "出刀总数", "剩余刀数", "QQ号", "群名片"]
    sheet.append(headers)
    for row in rows:
        member = row["member"]
        sheet.append([
            member.name,
            member.server_area or "",
            row["d1"],
            row["d2"],
            row["total"],
            row["remaining"],
            member.qq or "",
            member.group_card or "",
        ])
    style_header(sheet)
    autosize(sheet)


def build_day_attendance_sheet(sheet: Any, rows: list[dict[str, Any]]) -> None:
    headers = [
        "成员ID", "区服", "出刀数", "第一刀", "第二刀", "第三刀",
        "总伤害", "伤害占比", "剩余刀数", "QQ号", "群名片", "备注",
    ]
    sheet.append(headers)
    for row in rows:
        member = row["member"]
        damages = list(row["damages"])[:MAX_ATTACKS_PER_MEMBER]
        while len(damages) < MAX_ATTACKS_PER_MEMBER:
            damages.append(0)
        sheet.append([
            member.name,
            member.server_area or "",
            row["attack_count"],
            *damages,
            row["total_damage"],
            row["damage_share"],
            row["remaining"],
            member.qq or "",
            member.group_card or "",
            row["notes"],
        ])
    style_header(sheet)
    for row in sheet.iter_rows(min_row=2, min_col=8, max_col=8):
        row[0].number_format = "0.00%"
    autosize(sheet, max_width=48)


def build_note_sheet(sheet: Any, note: str) -> None:
    sheet.append(["项目", "说明"])
    sheet.append(["导出时间", now_text()])
    sheet.append(["模板状态", note])
    style_header(sheet)
    sheet.column_dimensions["A"].width = 16
    sheet.column_dimensions["B"].width = 100
    sheet["B3"].alignment = Alignment(wrap_text=True, vertical="top")


def build_damage_sheet(sheet: Any, attacks: list[dict[str, Any]]) -> None:
    by_member: dict[str, dict[str, Any]] = {}
    for attack in reversed(attacks):
        name = str(attack["member_name"])
        item = by_member.setdefault(
            name,
            {"name": name, "damages": [], "total": 0, "notes": []},
        )
        damage = int(attack["damage"] or 0)
        item["damages"].append(damage)
        item["total"] += damage
        note_parts = [
            str(attack.get("boss_label") or attack.get("boss_name") or "").strip(),
            str(attack.get("team_text") or "").strip(),
            "尾刀" if int(attack.get("is_tail") or 0) else "",
        ]
        note = " ".join(part for part in note_parts if part)
        if note:
            item["notes"].append(note)
    ranked = sorted(by_member.values(), key=lambda item: item["total"], reverse=True)
    total_damage = sum(int(item["total"]) for item in ranked)
    sheet.append(["排名", "成员ID", "第一刀", "第二刀", "第三刀", "总伤害", "伤害占比", "备注"])
    for index, item in enumerate(ranked, start=1):
        damages = list(item["damages"])[:3]
        while len(damages) < 3:
            damages.append(0)
        ratio = item["total"] / total_damage if total_damage else 0
        sheet.append([
            index,
            item["name"],
            *damages,
            item["total"],
            ratio,
            "；".join(item["notes"][:3]),
        ])
    style_header(sheet)
    for row in sheet.iter_rows(min_row=2, min_col=7, max_col=7):
        row[0].number_format = "0.00%"
    autosize(sheet)


def build_reference_damage_sheet(sheet: Any, attacks: list[dict[str, Any]], day_text: str) -> None:
    by_member: dict[str, dict[str, Any]] = {}
    for attack in reversed(attacks):
        name = str(attack["member_name"])
        item = by_member.setdefault(
            name,
            {"name": name, "damages": [], "total": 0, "notes": []},
        )
        damage = int(attack["damage"] or 0)
        item["damages"].append(damage)
        item["total"] += damage
        note_parts = [
            str(attack.get("boss_label") or attack.get("boss_name") or "").strip(),
            str(attack.get("team_text") or "").strip(),
            "尾刀" if int(attack.get("is_tail") or 0) else "",
        ]
        note = " ".join(part for part in note_parts if part)
        if note:
            item["notes"].append(note)
    ranked = sorted(by_member.values(), key=lambda item: item["total"], reverse=True)
    total_damage = sum(int(item["total"]) for item in ranked)
    average_damage = int(total_damage / len(ranked)) if ranked else 0
    sheet.append([f"联盟突袭 {day_text} 伤害统计", "均伤", "", average_damage, "", "", "", ""])
    sheet.append(["排名", "成员ID", "第一刀", "第二刀", "第三刀", "总伤害", "伤害占比", "备注"])
    for index, item in enumerate(ranked, start=1):
        damages = list(item["damages"])[:MAX_ATTACKS_PER_MEMBER]
        while len(damages) < MAX_ATTACKS_PER_MEMBER:
            damages.append(0)
        ratio = item["total"] / total_damage if total_damage else 0
        sheet.append([
            index,
            item["name"],
            *damages,
            item["total"],
            ratio,
            "；".join(item["notes"][:MAX_ATTACKS_PER_MEMBER]),
        ])
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["B1"].font = Font(bold=True)
    style_header_row(sheet, 2)
    for row in sheet.iter_rows(min_row=3, min_col=7, max_col=7):
        row[0].number_format = "0.00%"
    autosize(sheet)


def build_attack_detail_sheet(sheet: Any, attacks: list[dict[str, Any]]) -> None:
    sheet.append([
        "日期", "Day", "成员", "第几刀", "Boss", "Boss名", "伤害",
        "阵容", "尾刀", "备注", "来源", "记录时间",
    ])
    for attack in reversed(attacks):
        sheet.append([
            attack["battle_date"],
            attack.get("raid_day") or "",
            attack["member_name"],
            attack["attack_no"],
            attack.get("boss_label") or "",
            attack.get("boss_name") or "",
            int(attack["damage"] or 0),
            attack.get("team_text") or "",
            "是" if int(attack.get("is_tail") or 0) else "",
            attack.get("note") or "",
            attack.get("source") or "",
            attack.get("created_at") or "",
        ])
    style_header(sheet)
    autosize(sheet)


def build_snapshot_sheet(sheet: Any, snapshots: list[BossSnapshot]) -> None:
    sheet.append(["采样时间", "Day", "Boss", "Boss名", "等级", "血量", "截图", "备注"])
    for snapshot in reversed(snapshots):
        sheet.append([
            snapshot.captured_at,
            snapshot.raid_day or "",
            snapshot.boss_label or snapshot.boss_index or "",
            snapshot.boss_name,
            snapshot.level or "",
            format_hp(snapshot.current_hp, snapshot.total_hp, snapshot.percent),
            snapshot.image_path,
            snapshot.note,
        ])
    style_header(sheet)
    autosize(sheet, max_width=58)


def build_member_sheet(sheet: Any, members: list[Any]) -> None:
    sheet.append(["ID", "成员ID", "区服", "QQ号", "群名片", "启用"])
    for member in members:
        sheet.append([
            member.id,
            member.name,
            member.server_area or "",
            member.qq or "",
            member.group_card or "",
            "是" if member.active else "否",
        ])
    style_header(sheet)
    autosize(sheet)


def copy_export_to(path: Path, target_dir: Path | str) -> Path:
    target_root = Path(target_dir)
    target_root.mkdir(parents=True, exist_ok=True)
    target = target_root / path.name
    shutil.copy2(path, target)
    return target.resolve()
