from __future__ import annotations

import argparse
import csv
import io
import html
import sys
import urllib.parse
from email.parser import BytesParser
from email.policy import default
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from typing import Any

from openpyxl import load_workbook

from .core import GuildWarBot


DEFAULT_HOST = "127.0.0.1"
DEFAULT_PORT = 8790


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        prog="python -m guild_war_bot.admin_web",
        description="公会战机器人本地成员管理页面",
    )
    parser.add_argument("--host", default=DEFAULT_HOST)
    parser.add_argument("--port", type=int, default=DEFAULT_PORT)
    args = parser.parse_args(argv)

    bot = GuildWarBot()
    server = ThreadingHTTPServer((args.host, args.port), build_handler(bot))
    print(f"成员管理页面：http://{args.host}:{args.port}/")
    print("按 Ctrl+C 停止。")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print()
    finally:
        server.server_close()
        bot.close()
    return 0


def build_handler(bot: GuildWarBot) -> type[BaseHTTPRequestHandler]:
    class AdminHandler(BaseHTTPRequestHandler):
        def do_GET(self) -> None:
            path = self.path.split("?", 1)[0]
            if path == "/members.csv":
                self.send_csv(bot.export_members_csv())
                return
            if path != "/":
                self.send_response(404)
                self.end_headers()
                return
            query = urllib.parse.parse_qs(urllib.parse.urlparse(self.path).query)
            notice = (query.get("notice") or [""])[0]
            self.send_html(render_page(bot, notice))

        def do_POST(self) -> None:
            length = int(self.headers.get("Content-Length", "0"))
            body = self.rfile.read(length)
            content_type = self.headers.get("Content-Type", "")
            if content_type.startswith("multipart/form-data"):
                form = parse_multipart_form(content_type, body)
            else:
                raw = body.decode("utf-8", errors="replace")
                form = urllib.parse.parse_qs(raw)
            action = value(form, "action")

            if action == "add":
                notice = bot.add_member(
                    value(form, "name"),
                    value(form, "qq") or None,
                    value(form, "group_card") or None,
                )
            elif action == "update":
                notice = bot.update_member(
                    int(value(form, "id") or "0"),
                    value(form, "name"),
                    value(form, "qq") or None,
                    value(form, "active") == "1",
                    value(form, "group_card") or None,
                    value(form, "server_area") or None,
                )
            elif action == "bulk_update":
                delete_id = value(form, "delete_id")
                if delete_id:
                    notice = bot.delete_member(int(delete_id or "0"))
                else:
                    notice = bot.bulk_update_members(member_rows_from_form(form))
            elif action == "delete":
                notice = bot.delete_member(int(value(form, "id") or "0"))
            elif action == "import_csv":
                notice = bot.import_members_csv_text(value(form, "csv_text"))
            elif action == "upload_members":
                notice = import_uploaded_members(bot, form)
            else:
                notice = "未知操作。"

            self.redirect(f"/?notice={urllib.parse.quote(notice)}")

        def send_html(self, content: str) -> None:
            body = content.encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        def send_csv(self, content: str) -> None:
            body = ("\ufeff" + content).encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/csv; charset=utf-8")
            self.send_header(
                "Content-Disposition",
                'attachment; filename="guild-war-members.csv"',
            )
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        def redirect(self, location: str) -> None:
            self.send_response(303)
            self.send_header("Location", location)
            self.end_headers()

        def log_message(self, fmt: str, *args: Any) -> None:
            print(f"[Admin] {self.address_string()} - {fmt % args}")

    return AdminHandler


def value(form: dict[str, list[str]], key: str) -> str:
    return (form.get(key) or [""])[0].strip()


def member_rows_from_form(form: dict[str, list[str]]) -> list[dict[str, str]]:
    ids = form.get("id") or []
    rows: list[dict[str, str]] = []
    for index, member_id in enumerate(ids):
        rows.append(
            {
                "id": member_id,
                "name": indexed_value(form, "name", index),
                "server_area": indexed_value(form, "server_area", index),
                "qq": indexed_value(form, "qq", index),
                "group_card": indexed_value(form, "group_card", index),
                "active": indexed_value(form, "active", index) or "0",
            }
        )
    return rows


def indexed_value(form: dict[str, list[str]], key: str, index: int) -> str:
    values = form.get(key) or []
    if index >= len(values):
        return ""
    return values[index].strip()


def parse_multipart_form(content_type: str, body: bytes) -> dict[str, list[str]]:
    raw_message = (
        f"Content-Type: {content_type}\r\n"
        "MIME-Version: 1.0\r\n\r\n"
    ).encode("utf-8") + body
    message = BytesParser(policy=default).parsebytes(raw_message)
    form: dict[str, list[str]] = {}
    for part in message.iter_parts():
        name = part.get_param("name", header="content-disposition")
        if not name:
            continue
        filename = part.get_filename()
        payload = part.get_payload(decode=True) or b""
        if filename:
            form.setdefault(name, []).append(filename)
            form.setdefault(f"{name}__bytes", []).append(payload.decode("latin1"))
        else:
            charset = part.get_content_charset() or "utf-8"
            form.setdefault(name, []).append(payload.decode(charset, errors="replace"))
    return form


def import_uploaded_members(bot: GuildWarBot, form: dict[str, list[str]]) -> str:
    filename = value(form, "members_file")
    raw_payload = (form.get("members_file__bytes") or [""])[0].encode("latin1")
    if not filename or not raw_payload:
        return "请先选择 CSV 或 XLSX 文件。"
    suffix = filename.rsplit(".", 1)[-1].lower()
    if suffix == "csv":
        csv_text = raw_payload.decode("utf-8-sig", errors="replace")
    elif suffix == "xlsx":
        csv_text = xlsx_bytes_to_csv_text(raw_payload)
    else:
        return "只支持上传 .csv 或 .xlsx 文件。"
    return bot.import_members_csv_text(csv_text)


def xlsx_bytes_to_csv_text(payload: bytes) -> str:
    workbook = load_workbook(io.BytesIO(payload), data_only=True)
    sheet = workbook["机器人导入"] if "机器人导入" in workbook.sheetnames else workbook.active
    detail_area_by_name = read_detail_area_by_name(workbook)
    output = io.StringIO()
    writer = csv.writer(output)
    headers: list[str] | None = None
    for row in sheet.iter_rows(values_only=True):
        values = ["" if value is None else str(value).strip() for value in row]
        if any(values):
            if headers is None:
                headers = values
                if "server_area" not in headers and "区服" not in headers:
                    headers = [*headers, "server_area"]
                writer.writerow(headers)
                continue
            if headers and len(values) < len(headers):
                values = [*values, *([""] * (len(headers) - len(values)))]
            if headers and "server_area" in headers:
                name_index = headers.index("name") if "name" in headers else -1
                area_index = headers.index("server_area")
                if name_index >= 0 and not values[area_index]:
                    values[area_index] = detail_area_by_name.get(values[name_index], "")
            writer.writerow(values)
    return output.getvalue()


def read_detail_area_by_name(workbook: Any) -> dict[str, str]:
    if "成员明细" not in workbook.sheetnames:
        return {}
    sheet = workbook["成员明细"]
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = ["" if value is None else str(value).strip() for value in header_row]
    if "name" not in headers or "server_area" not in headers:
        return {}
    name_index = headers.index("name")
    area_index = headers.index("server_area")
    result: dict[str, str] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        values = ["" if value is None else str(value).strip() for value in row]
        if len(values) <= max(name_index, area_index):
            continue
        if values[name_index] and values[area_index]:
            result[values[name_index]] = values[area_index]
    return result


def render_page(bot: GuildWarBot, notice: str = "") -> str:
    records = bot.list_member_records(include_inactive=True)
    active_count = sum(1 for record in records if record.active)
    rows = "\n".join(render_member_row(record) for record in records)
    tabs = render_area_tabs(records)
    notice_html = (
        f'<div class="notice">{escape(notice)}</div>'
        if notice
        else ""
    )
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>公会战成员管理</title>
  <style>
    :root {{
      color-scheme: light;
      --bg: #f6f7f9;
      --panel: #ffffff;
      --line: #d9dee7;
      --text: #182033;
      --muted: #667085;
      --primary: #1664d9;
      --danger: #c93434;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      background: var(--bg);
      color: var(--text);
      font-family: "Microsoft YaHei", "Segoe UI", sans-serif;
      font-size: 14px;
    }}
    main {{
      max-width: 1080px;
      margin: 0 auto;
      padding: 24px;
    }}
    header {{
      display: flex;
      align-items: end;
      justify-content: space-between;
      gap: 16px;
      margin-bottom: 18px;
    }}
    h1 {{
      margin: 0 0 6px;
      font-size: 24px;
      font-weight: 700;
    }}
    .sub {{ color: var(--muted); }}
    section {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 16px;
      margin-bottom: 16px;
    }}
    .notice {{
      border: 1px solid #a8d3b2;
      background: #edf8ef;
      color: #226331;
      border-radius: 6px;
      padding: 10px 12px;
      margin-bottom: 16px;
    }}
    form.add {{
      display: grid;
      grid-template-columns: minmax(140px, 1fr) minmax(170px, 1fr) minmax(190px, 1fr) auto;
      gap: 10px;
      align-items: end;
    }}
    label {{
      display: grid;
      gap: 6px;
      color: var(--muted);
      font-size: 12px;
    }}
    input, select {{
      width: 100%;
      min-height: 36px;
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 7px 9px;
      color: var(--text);
      background: #fff;
      font: inherit;
    }}
    textarea {{
      width: 100%;
      min-height: 160px;
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 9px;
      color: var(--text);
      background: #fff;
      font: 13px/1.5 Consolas, "Microsoft YaHei", monospace;
      resize: vertical;
    }}
    button {{
      min-height: 36px;
      border: 1px solid transparent;
      border-radius: 6px;
      padding: 7px 12px;
      background: var(--primary);
      color: #fff;
      font: inherit;
      cursor: pointer;
      white-space: nowrap;
    }}
    button.secondary {{
      background: #fff;
      color: var(--primary);
      border-color: var(--primary);
    }}
    button.danger {{
      background: #fff;
      color: var(--danger);
      border-color: #e4b4b4;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
    }}
    th, td {{
      border-bottom: 1px solid var(--line);
      padding: 10px 8px;
      text-align: left;
      vertical-align: middle;
    }}
    th {{
      color: var(--muted);
      font-size: 12px;
      font-weight: 600;
      background: #fafbfc;
    }}
    tr.inactive td, .member-form.inactive {{
      color: #98a2b3;
      background: #fafafa;
    }}
    .member-form {{
      display: grid;
      grid-template-columns: 56px minmax(110px, .8fr) minmax(140px, 1fr) minmax(130px, .9fr) minmax(170px, 1.2fr) 96px 92px;
      gap: 8px;
      align-items: center;
    }}
    .member-header {{
      color: var(--muted);
      font-size: 12px;
      font-weight: 600;
      background: #fafbfc;
      border-bottom: 1px solid var(--line);
      padding: 8px;
    }}
    .empty {{
      color: var(--muted);
      text-align: center;
      padding: 24px 0;
    }}
    .tools {{
      display: grid;
      grid-template-columns: 1fr 1.4fr;
      gap: 16px;
      align-items: start;
    }}
    .hint {{
      color: var(--muted);
      font-size: 13px;
      line-height: 1.7;
      margin: 8px 0 12px;
    }}
    .sample {{
      display: block;
      color: var(--muted);
      background: #f7f8fa;
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 8px;
      overflow: auto;
    }}
    .tabs {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      margin-bottom: 12px;
    }}
    .tab {{
      border: 1px solid var(--line);
      border-radius: 999px;
      padding: 7px 12px;
      background: #fff;
      color: var(--text);
      cursor: pointer;
    }}
    .tab.active {{
      background: var(--primary);
      border-color: var(--primary);
      color: #fff;
    }}
    .row-actions {{
      display: flex;
      gap: 8px;
      align-items: center;
      margin: 12px 0;
    }}
    @media (max-width: 760px) {{
      main {{ padding: 14px; }}
      header {{ display: block; }}
      form.add, .member-form, .tools {{ grid-template-columns: 1fr; }}
      table, thead, tbody, tr, td {{ display: block; width: 100%; }}
      thead {{ display: none; }}
      td {{ border-bottom: 0; padding: 6px 0; }}
      tr {{ border-bottom: 1px solid var(--line); padding: 10px 0; }}
    }}
  </style>
</head>
<body>
<main>
  <header>
    <div>
      <h1>公会战成员管理</h1>
      <div class="sub">已启用 {active_count} 人 / 共 {len(records)} 人，QQ 号用于自动识别，Q群备注名用于提醒和人工核对。</div>
    </div>
  </header>
  {notice_html}
  <section>
    <form class="add" method="post">
      <input type="hidden" name="action" value="add">
      <label>游戏名 / 成员名<input name="name" required placeholder="例如：张三"></label>
      <label>区服<input name="server_area" placeholder="例如：Q区 / V区"></label>
      <label>QQ 号<input name="qq" inputmode="numeric" placeholder="例如：1255348850"></label>
      <label>Q群备注名<input name="group_card" placeholder="例如：［Q区］张三"></label>
      <button type="submit">新增成员</button>
    </form>
  </section>
  <section class="tools">
    <div>
      <h2>批量导出</h2>
      <p class="hint">导出当前全部成员，包含停用成员。字段包含 <code>group_card</code>，可用于保存 Q群备注名。</p>
      <a href="/members.csv"><button type="button">导出成员 CSV</button></a>
    </div>
    <form method="post" enctype="multipart/form-data">
      <input type="hidden" name="action" value="upload_members">
      <h2>上传表格导入 / 更新</h2>
      <p class="hint">支持 <code>.xlsx</code> 和 <code>.csv</code>。XLSX 优先读取 <code>机器人导入</code> sheet。支持列：<code>id,name,qq,group_card,active</code>，也兼容 <code>Q群备注名</code>。</p>
      <label>选择文件<input type="file" name="members_file" accept=".xlsx,.csv" required></label>
      <button type="submit" style="margin-top:10px;">上传并导入</button>
    </form>
    <form method="post">
      <input type="hidden" name="action" value="import_csv">
      <h2>粘贴 CSV 导入 / 更新</h2>
      <p class="hint">有 ID 时按 ID 更新；没有 ID 时按成员名更新；都不存在则新增。<code>active</code> 填 1 启用、0 停用。</p>
      <small class="sample">id,name,qq,group_card,active
1,张三,1255348850,［Q区］张三,1
,新成员,123456789,［Q区］新成员,1</small>
      <label style="margin-top:10px;">CSV 内容<textarea name="csv_text" required placeholder="把 CSV 内容粘贴到这里"></textarea></label>
      <button type="submit" style="margin-top:10px;">导入 / 更新成员</button>
    </form>
  </section>
  <section>
    <h2>成员列表</h2>
    {tabs}
    <form method="post">
      <input type="hidden" name="action" value="bulk_update">
      <div class="member-form member-header">
        <div>ID</div><div>区服</div><div>游戏名</div><div>QQ号</div><div>Q群备注名</div><div>状态</div><div></div>
      </div>
      {rows if rows else '<div class="empty">暂无成员</div>'}
      <div class="row-actions">
        <button type="submit">保存全部修改</button>
        <span class="hint">修改多行后点一次即可保存全部。</span>
      </div>
    </form>
  </section>
  <script>
    function showArea(area) {{
      document.querySelectorAll('.tab').forEach(tab => tab.classList.toggle('active', tab.dataset.area === area));
      document.querySelectorAll('[data-member-area]').forEach(row => {{
        row.style.display = area === '全部' || row.dataset.memberArea === area ? '' : 'none';
      }});
    }}
    showArea('全部');
  </script>
</main>
</body>
</html>"""


def render_area_tabs(records: list[Any]) -> str:
    counts: dict[str, int] = {"全部": len(records)}
    for record in records:
        area = record.server_area or "未分区"
        counts[area] = counts.get(area, 0) + 1
    preferred = ["全部", "Q区", "V区", "未分区"]
    areas = [area for area in preferred if area in counts]
    areas.extend(sorted(area for area in counts if area not in preferred))
    buttons = [
        f'<button class="tab" type="button" data-area="{escape(area)}" onclick="showArea(\'{escape(area)}\')">{escape(area)}（{counts[area]}）</button>'
        for area in areas
    ]
    return '<div class="tabs">' + "\n".join(buttons) + "</div>"


def render_member_row(record: Any) -> str:
    active_selected = "selected" if record.active else ""
    inactive_selected = "" if record.active else "selected"
    row_class = "" if record.active else " inactive"
    area = record.server_area or "未分区"
    return f"""<div class="member-form{row_class}" data-member-area="{escape(area)}">
      <div>#{record.id}</div>
      <input type="hidden" name="id" value="{record.id}">
      <input name="server_area" value="{escape(record.server_area or '')}" placeholder="Q区 / V区">
      <input name="name" required value="{escape(record.name)}">
      <input name="qq" inputmode="numeric" value="{escape(record.qq or '')}" placeholder="QQ 号">
      <input name="group_card" value="{escape(record.group_card or '')}" placeholder="Q群备注名">
      <select name="active">
        <option value="1" {active_selected}>启用</option>
        <option value="0" {inactive_selected}>停用</option>
      </select>
      <button class="danger" type="submit" name="delete_id" value="{record.id}" onclick="return confirm('确定删除这个成员和他的出刀记录？');">删除</button>
</div>"""


def escape(value: str) -> str:
    return html.escape(value, quote=True)


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
