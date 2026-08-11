from __future__ import annotations

import sqlite3
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from guild_war_bot.union_suite import UnionRaidSuite


class UnionRaidSuiteTests(unittest.TestCase):
    def test_records_snapshot_attack_and_exports_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            suite = UnionRaidSuite(
                db_path=root / "guild_war.db",
                sample_db_path=root / "union_sample.db",
            )
            try:
                suite.add_member("DORO", "Q区", "123", "〔Q区〕DORO")
                member = suite.list_members()[0]
                self.assertIn("Boss 进度已保存", suite.add_boss_snapshot(
                    raid_day=1,
                    boss_index=1,
                    boss_label="I",
                    boss_name="殓巾",
                    level=10,
                    current_hp=100,
                    total_hp=200,
                    note="test",
                ))
                self.assertIn("已记录", suite.add_attack(
                    member_id=member.id,
                    battle_date_text="2026-06-12",
                    damage=123456,
                    note="",
                    boss_label="I",
                    boss_name="殓巾",
                    team_text="丽塔 / 皇冠",
                    raid_day=1,
                    is_tail=True,
                ))
                output = suite.export_workbook(
                    template_path=root / "missing-template.xlsx",
                    output_dir=root / "exports",
                )
            finally:
                suite.close()

            self.assertTrue(output.exists())
            workbook = load_workbook(output, data_only=True)
            self.assertIn("当前-考勤", workbook.sheetnames)
            self.assertIn("当前-统伤", workbook.sheetnames)
            self.assertEqual(workbook["当前-统伤"]["B2"].value, "DORO")
            self.assertEqual(workbook["出刀明细"]["G2"].value, 123456)
            self.assertEqual(workbook["Boss快照"]["D2"].value, "殓巾")

    def test_import_sampled_members_from_existing_sample_db(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            sample_db = root / "union_sample.db"
            conn = sqlite3.connect(sample_db)
            conn.execute(
                """
                create table member_samples (
                    name text,
                    power integer,
                    level integer,
                    online_text text,
                    online_status text,
                    updated_at text
                )
                """
            )
            conn.execute(
                "insert into member_samples values (?, ?, ?, ?, ?, ?)",
                ("NING", 999, 500, "1分钟前", "在线", "2026-06-12 12:00:00"),
            )
            conn.commit()
            conn.close()

            suite = UnionRaidSuite(
                db_path=root / "guild_war.db",
                sample_db_path=sample_db,
            )
            try:
                notice = suite.import_sampled_members(limit=32, server_area="Q区")
                members = suite.list_members()
            finally:
                suite.close()

            self.assertIn("新增 1 人", notice)
            self.assertEqual(members[0].name, "NING")
            self.assertEqual(members[0].server_area, "Q区")


if __name__ == "__main__":
    unittest.main()
